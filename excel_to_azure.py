import openpyxl
from keys import ConnectionString

# load data from excel
wb = openpyxl.load_workbook("krypto.xlsx")
ws = wb.active

last_row = 5
value = 1

while value:
    last_row +=1
    value = ws['A' + str(last_row)].value

dates = [ws['A' + str(i)].value for i in range(2, last_row)]

eur = [ws['B' + str(i)].value for i in range(2, last_row)]
eureth = [ws['C' + str(i)].value for i in range(2, last_row)]
eurbtc = [ws['D' + str(i)].value for i in range(2, last_row)]

from azure.cosmosdb.table import TableService

tableName = "balances"

tableServiceClient = TableService(endpoint_suffix = "table.cosmos.azure.com", connection_string= ConnectionString)
if not tableServiceClient.exists(tableName):
    print()
    print("The table does not exist")
    exit()

for i in range(2, last_row):
    day = ws['A' + str(i)].value
    eur = ws['B' + str(i)].value
    eur_eth = ws['C' + str(i)].value
    eur_btc = ws['D' + str(i)].value
    eth = ws['E' + str(i)].value
    btc = ws['F' + str(i)].value

    entity = {
        "PartitionKey": str(day.year*100 + day.month),
        "RowKey": str(day.strftime("%Y%m%d")),
        "EUR": eur,
        "EURETH": eur_eth,
        "EURBTC": eur_btc,
        "ETH": eth,
        "BTC": btc
    }

    tableServiceClient.insert_or_replace_entity(tableName, entity, timeout=None)
