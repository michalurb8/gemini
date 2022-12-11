from matplotlib.axis import YAxis
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

import openpyxl
from datetime import date, timedelta, datetime

# save that data to excel:
wb = openpyxl.load_workbook("krypto.xlsx")
ws = wb.active

last_row = 5
value = 1

while value:
    last_row +=1
    value = ws['A' + str(last_row)].value

dates = [ws['A' + str(i)].value for i in range(2, last_row)]

eur = [ws['B' + str(i)].value for i in range(2, last_row)]
eureth = [ws['C' + str(i)].value for i in range(2, last_row)]
eurbtc = [ws['D' + str(i)].value for i in range(2, last_row)]


yrange = int(1.1*(max(max(eurbtc) + max(eureth), max(eur))))
step = 10

plt.figure(figsize=(15,8), dpi=100)
ax = plt.gca()
plt.margins(x=0)
plt.stackplot(dates, eurbtc, eureth, eur, labels = ['EUR/BTC', 'EUR/ETH', 'EUR'],colors=[(0.92,0.5,0.25), (0.32,0.48,0.80), (0.3,0.3,0.3)])

xticksep = 4
plt.xticks([date.today() - timedelta(days=x) for x in range(0, (datetime.today() - dates[0]).days, xticksep)])
for index, day in enumerate(dates):
    plt.vlines(day, 0, (eureth[index] + eurbtc[index] + eur[index]), color=(0,0,0,0.1), linewidth=1)
    plt.scatter(day, (eureth[index] + eurbtc[index] + eur[index]), c='black', s=8)
plt.axvline(date(2022,9,15), color=(1,0,0,0.3), linewidth=2)

plt.grid(axis='y',alpha=0.3,color='BLACK')
plt.yticks(range(step, yrange+step, step), size=12)
ax.xaxis.set_major_locator(mdates.DayLocator(bymonthday=[5,10,15,20,25]))
ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m.%Y'))
ax.tick_params(axis='x', which='major', width = 1, length = 3, labelsize = 10, colors='black', labelrotation= 90)

ax.xaxis.set_minor_locator(mdates.MonthLocator())
ax.xaxis.set_minor_formatter(mdates.DateFormatter('%b'))
ax.tick_params(axis='x', which='minor', width = 2, length = 7, labelsize = 20, colors=(0.5,0,0), labelrotation= 80)

ax.yaxis.set_major_formatter('€{x:1.2f}')
ax.tick_params(axis='y', labelrotation= 0)


plt.legend(fontsize=15)
plt.tight_layout()
plt.show()