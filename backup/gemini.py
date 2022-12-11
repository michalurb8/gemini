import hmac
import requests
import json
import base64
import hmac
import hashlib
import time, datetime

import openpyxl
from keys import url, gemini_api_key, gemini_api_secret

nonce = time.time()

payload = {
    "request": "/v1/notionalbalances/eur",
    "nonce": nonce
}

encoded_payload = json.dumps(payload).encode()
b64 = base64.b64encode(encoded_payload)
signature = hmac.new(gemini_api_secret, b64, hashlib.sha384).hexdigest()

request_headers = {
    'Content-Type': "text/plain",
    'Content-Length': "0",
    'X-GEMINI-APIKEY': gemini_api_key,
    'X-GEMINI-PAYLOAD': b64,
    'X-GEMINI-SIGNATURE': signature,
    'Cache-Control': "no-cache"
}

response = requests.post(url = url, headers = request_headers).json()
values = {}
for res in response:
    amount = float(res['amount'])
    notional = round(100*float(res['amountNotional']))/100.0
    values[res['currency']] = (amount, notional)

eur = values['EUR'][1]
eureth = values['ETH'][1]
eurbtc = values['BTC'][1]
eth = values['ETH'][0]
btc = values['BTC'][0]

array = [eur, eureth, eurbtc, eth, btc]




# save that data to excel:
wb = openpyxl.load_workbook("krypto.xlsx")
ws = wb.active

row = 5
value = 1

while value:
    row +=1
    value = ws['A' + str(row)].value

if (datetime.datetime.today() - (ws['A' + str(row-1)].value)).days == 0:
    row -= 1

date_cell = ws['A' + str(row)]
date_cell.value = datetime.date.today()
date_cell.number_format = "DD.MM.YYYY"

ws['B' + str(row)].value = eur
ws['C' + str(row)].value = eureth
ws['D' + str(row)].value = eurbtc
ws['E' + str(row)].value = eth
ws['F' + str(row)].value = btc

wb.save("krypto.xlsx")